import smtplib
import tkinter as tk
from tkinter import ttk
import numpy as np
# from PIL import ImageTk,Image
import re
import sqlite3 as sql
import cv2 as cv
import math
from tkinter import filedialog
import os
from imea.measure_2d import statistical_length
from skimage import io, color
import openpyxl
# import pandas as pd
# import keyboard
from color_match import img_color_match
import matplotlib.pyplot as plt
#import Rect_identifier

za=[0,0,0,0,0,0,0,0,0]
anan=14
clicks = 0
Large_Font=("Verdana",12)
Normal_Font=("Verdana",10)
Small_Font=("Verdana",8)
title="Body Plan"
exchange="Cb1"
DatCounter=9000
programName="cb1"
chartLoad=True
canom=1
zeksen=0
cnt=0
key="alsşdasdlsşlilişlgişlşi"

sayac99=0
sayac991=0

class warningPage(tk.Frame):
    def __init__(self,parent,controller):
        tk.Frame.__init__(self,parent)
        global basic_locker_app_35
        global sq
        global menubar
        global filemenu
        global disable_menu
        global enable_menu
        while True:
            disable_menu()
            break

        def basic_locker_app_35():
            nick=e.get()
            password=b.get()

            logs=sql.connect("logs.sqlite")
            im=logs.cursor()
            logs.commit()
            im.execute("""SELECT isim,şifre FROM kullanıcılar""")

            veriler=im.fetchall()
            im.close()
            while True :
                global cnt
                global eec
                if cnt<1:
                    global eec
                    if (nick,password) in veriler:
                        command=lambda:controller.show_frame(composite)
                        command()
                        enable_menu()
                        e.delete(0,"end")
                        b.delete(0,"end")
                        break
                    else:
                        cnt=cnt+1
                        eec=tk.Label(self,text="Wrong nickname or password ! Please Try again!",font=Small_Font)
                        eec.place(x=800,y=335)
                        break
                else:
                    eec.destroy()
                    if (nick,password) in veriler:
                        enable_menu()
                        cnt=0
                        eec.destroy()
                        e.delete(0,"end")
                        b.delete(0,"end")
                        command=lambda:controller.show_frame(composite)
                        command()
                        break
                    else:
                        eec=tk.Label(self,text="Wrong nickname or password ! Please Try again!",font=Small_Font)
                        eec.place(x=800,y=335)
                        break

        e=ttk.Entry(self)
        e.place(x=910,y=275)
        b=ttk.Entry(self,show="*")
        b.place(x=910,y=310)
        label1=tk.Label(self,text="Welcome to The SBCB!",font=Large_Font)
        label1.pack(padx=20,pady=20)
        label12=tk.Label(self,text="If you have an account please log in.",font=Large_Font)
        label12.pack(padx=20,pady=20)
        labelName=ttk.Label(self,text="Nickname:",font=Normal_Font,state="normal")
        labelName.place(x=830,y=275)
        password=ttk.Label(self,text="Password:",font=Normal_Font)
        password.place(x=830,y=310)
        button31=ttk.Button(self,text="Log-in",command=basic_locker_app_35)
        button31.place(x=930,y=355)
        button54=ttk.Button(self,text="Register", command=lambda:controller.show_frame(registerPage))
        button54.place(x=930.5,y=410)
        copyrights= tk.Label(self,text="© Copyright of this software has taken by D.C.!D.C. Is MORE.",font=Small_Font)
        copyrights.place(x=1550,y=650)
class registerPage(tk.Frame):
    def __init__(self,parent,controller):
        tk.Frame.__init__(self,parent)
        copyrights= tk.Label(self,text="© Copyright of this software has taken by D.C.!D.C. Is MORE.",font=Small_Font)
        copyrights.place(x=1550,y=650)
        global uyarı
        global uyarı2
        global uyarı4
        global uyarı5
        global enable_menu
        global logout
        xy=ttk.Entry(self)
        xy.place(x=930,y=200)

        yx=ttk.Entry(self,show="*")
        yx.place(x=930,y=245)


        def show():
            global sayac99

            abo=yx.get()
            sayac99+=1


            while True:
                if sayac99==1:
                    yx.delete(0,"end")
                    yx.configure(text="",show="")
                    yx.insert(END,abo)

                    break
                else:
                    yx.delete(0,"end")
                    yx.configure(text="",show="*")
                    yx.insert(END,abo)
                    sayac99=0
                    break
        yxy=ttk.Entry(self,show='*')
        yxy.place(x=930,y=290)
        def show1():
            global sayac991
            abo1=yxy.get()
            sayac991+=1
            while True:
                if sayac991==1:
                    yxy.delete(0,"end")
                    yxy.configure(text="",show="")
                    yxy.insert(END,abo1)

                    break
                else:
                    yxy.delete(0,"end")
                    yxy.configure(text="",show="*")
                    yxy.insert(END,abo1)
                    sayac991=0
                    break

        label415=tk.Label(self,text="Please Enter Your New Nickname :",font=Small_Font)
        label415.place(x=710,y=200)

        label424=tk.Label(self,text="Please Enter Your New Password :",font=Small_Font)
        label424.place(x=710,y=245)

        label4241=tk.Label(self,text="Please Enter Your New Password Again! :",font=Small_Font)
        label4241.place(x=670,y=290)

        label1=tk.Label(self,text="Welcome to The SBCB!",font=Large_Font)
        label1.pack(padx=20,pady=20)

        label12=tk.Label(self,text="Please Register to The System!",font=Large_Font)
        label12.pack(padx=20,pady=20)





        uyarı=tk.Label(self,text="-Nickname or Password must be at least 6 character.",font=Small_Font)
        uyarı.place(x=800,y=490)

        uyarı3=tk.Label(self,text="-Nickname should not start with number",font=Small_Font)
        uyarı3.place(x=800,y=460)

        uyarı2=tk.Label(self,text="-Please enter the same Passwords in the blanks.",font=Small_Font)
        uyarı2.place(x=800,y=520)
        tıkla=0



        serial=ttk.Label(self,text="Please Enter Your Subscription Key :")
        serial.place(x=720,y=380)

        serial=ttk.Label(self,text="Please Enter Your Email Adress :")
        serial.place(x=740,y=335)

        mails61=ttk.Entry(self)
        mails61.place(x=930,y=335)
        uyarı4=tk.Label(self,text=("1"),font=Small_Font)

        anskey11=ttk.Entry(self)
        anskey11.place(x=930,y=380)

        book=ttk.Checkbutton(self,text="Show Paswoord",command=show)
        book.place(x=1075,y=245)

        book=ttk.Checkbutton(self,text="Show Paswoord",command=show1)
        book.place(x=1075,y=290)

        uyarı5=tk.Label(self,text=("Username has already exists!"),font="Verdana 10 underline")
        def tık():
            global key
            global uyarı4
            global uyarı5


            anskey=anskey11.get()
            c=xy.get()
            d=yx.get()
            ce=yxy.get()
            name = re.findall("^\d", c)
            logs=sql.connect("logs.sqlite")
            im=logs.cursor()
            im.execute("""SELECT isim FROM kullanıcılar""")
            logs.commit()
            isin=im.fetchall()
            im.close()
            isimler = []
            for t in isin:
                for x in t:
                    isimler.append(x)
            while True:
                if c in isimler:
                    uyarı4.place_forget()
                    uyarı.place(x=800,y=490)
                    uyarı2.place(x=800,y=520)
                    uyarı3.place(x=800,y=460)
                    uyarı5.place(x=1075,y=200)

                    break
                else:
                    uyarı5.place_forget()

                    if len(c)<=6 or len(d)<=6 or ce!=d or key!=anskey or len(name)!=0 :
                        uyarı4.place_forget()
                        uyarı.place(x=800,y=490)
                        uyarı2.place(x=800,y=520)
                        uyarı3.place(x=800,y=460)


                        break
                    else:
                        uyarı5.place_forget()
                        uyarı.place_forget()
                        uyarı2.place_forget()
                        uyarı3.place_forget()
                        uyarı4=tk.Label(self,text=("Well Come"+" "+c),font=Small_Font)
                        uyarı4.place(x=850,y=460)

                        logs=sql.connect("logs.sqlite")
                        im=logs.cursor()
                        im.execute("INSERT INTO kullanıcılar VALUES ('{}', '{}','{}')".format(c,d,key))
                        logs.commit()
                        im.close()
                        key="asdasdnbahtoıvjmovamvow>£#½$#£>{>{>£$>£#>£#>#£tatovemoaetoaw>>>>>>>>>2#12#>#33"
                        break
        mailss=tk.Label(self,text="This e-mail does not exist",font="Verdana 10 underline")
        def logout():
            global uyarı
            global uyarı2
            global disable_menu
            global uyarı4
            global uyarı5
            uyarı5.place_forget()

            mailss.place_forget()
            mails61.delete(0,"end")
            anskey11.delete(0,"end")
            uyarı.place(x=800,y=490)
            uyarı3.place(x=800,y=460)
            uyarı2.place(x=800,y=520)
            uyarı4.place_forget()
            xy.delete(0,"end")
            yx.delete(0,"end")
            yxy.delete(0,"end")
            command=lambda:controller.show_frame(warningPage)
            command()
        def keygen():
            global key
            count=1
            seq = "ABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890"
            for i in range(count):
                key=str('-'.join(''.join(random.choice(seq) for _ in range(5)) for _ in range(5)))

        regex= '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
        def mail():

            keygen()
            try:
                alan=str(mails61.get())
                if(re.search(regex,alan)):
                    mailss.place_forget()
                    content= key
                    print(content)
                    mail=smtplib.SMTP("smtp.gmail.com",587)
                    mail.ehlo()
                    mail.starttls()
                    mail.login("sbcbprogramme@gmail.com","Denizckr61::")
                    mail.sendmail("sbcbprogramme@gmail.com",alan,content)
                    mail.quit()
                else:
                    mailss.place(x=1180,y=333)

            except Exception as e:
                popupmsg(e)
        mailbut=ttk.Button(self,text=("Get Sub Key"),command=mail)
        mailbut.place(x=1080,y=333)

        button54=ttk.Button(self,text="Register",command=tık)
        button54.place(x=912.5,y=555)
        button55=ttk.Button(self,text="Log-in",command=logout)
        button55.place(x=912.5,y=600)
fotoolcer=0
fotoolcers=0
class composite(tk.Frame):
    def __init__(self,parent,controller):
        tk.Frame.__init__(self,parent)
        global fotoolcer
        global params1
        global params2
        global filename
        global basename1
        global label1
        global disable_menu
        global fotoolcers
        global refPt
        global label12
        global coefficientval
        global black
        global vf_iter_counter
        global pop_yes_or_no
        black=0
        vf_iter_counter=0


        def blacklivesmatter():
            global black
            black+=1
            while True:
                if black==1:
                    print(black)

                    break
                else:
                    black=0
                    print(black)
                    break
        def vf_iterr():
            global vf_iter_counter
            vf_iter_counter+=1
            while True:
                if vf_iter_counter==1:
                    print(vf_iter_counter)

                    break
                else:
                    vf_iter_counter=0
                    print(vf_iter_counter)
                    break
        def measurebar():
            global fotoolcer
            global label1
            global params1
            global params2
            global src
            global src1
            global basename1
            global coefficientval
            global radius
            global circles
            global x61
            global radiusus
            popup1=tk.Tk()
            popup1.geometry('400x300')
            hesaplistesi = ttk.Treeview(popup1)
            hesaplistesi['columns']=('Fiber No','Radius','Area')
            hesaplistesi.column('#0', width=0, stretch=NO)
            hesaplistesi.column('Fiber No', anchor=CENTER, width=80)
            hesaplistesi.column('Radius', anchor=CENTER, width=80)
            hesaplistesi.column('Area', anchor=CENTER, width=80)
            hesaplistesi.heading('#0', text='', anchor=CENTER)
            hesaplistesi.heading('Fiber No', text='Fiber No', anchor=CENTER)
            hesaplistesi.heading('Radius', text='Radius', anchor=CENTER)
            hesaplistesi.heading('Area', text='Area', anchor=CENTER)

            for i in range(x61):
                hesaplistesi.insert(parent='', index=i, iid=i, text='', values=(i+1,radiusus[i],alanlar[i]))
            hesaplistesi.pack()
            measures =tk.Text(popup1)
            measures.delete(1.0, "end")

            def kapat():
                popup1.destroy()
            popup1.wm_title("Measurement")

            B1=ttk.Button(popup1,text="Close",command =kapat)
            B1.pack()
            popup1.mainloop()
        refPt = []
        cropping = False
        refPt1 = []
        def uzunolcer():
            try:
                global params1
                global params2
                global basename1
                global refPt
                global refPt1
                global x1
                global x2
                global y1
                global y2
                global cropping
                global fotoolcers
                global label12
                global ölçü
                global coefficientval
                global radius
                global circles
                global alan
                global alanlar
                global radiusus

                def lengthreader(event, x, y, flags, param):
                	global refPt1, cropping
                	if event == cv.EVENT_LBUTTONDOWN:
                		refPt1 = [(x, y)]
                		cropping = True
                	elif event == cv.EVENT_LBUTTONUP:
                		refPt1.append((x, y))
                		cropping = False
                		cv.line(image, refPt1[0], refPt1[1], (0, 255, 0), 1)
                		cv.imshow("image", image)
                image = cv.imread("{}".format(basename1))
                clone = image.copy()
                cv.namedWindow("image")
                cv.setMouseCallback("image", lengthreader)
                # keep looping until the 'q' key is pressed
                while True:
                	# display the image and wait for a keypress
                	cv.imshow("image", image)
                	key = cv.waitKey(1) & 0xFF
                	# if the 'r' key is pressed, reset the cropping region
                	if key == ord("r"):
                		image = clone.copy()
                	# if the 'c' key is pressed, break from the loop
                	elif key == ord("c"):
                		break
                # if there are two reference points, then crop the region of interest
                # from teh image and display it
##                print(refPt1)
                xdeğerleri=[]
                ydeğerleri=[]
                for i in refPt1:
                    xdeğerleri.append(i[0])
                    ydeğerleri.append(i[1])
                for i in xdeğerleri:
                    x1=xdeğerleri[0]
                    x2=xdeğerleri[1]
                for i in ydeğerleri:
                    y1=ydeğerleri[0]
                    y2=ydeğerleri[1]

                a=((x1-x2)**2+(y1-y2)**2)**(1/2)
                try:
                    ölçü=float(anskey1133.get())
                except Exception as e:
                    popupmsg("Enter a valid value")
                coefficientval=float(ölçü/a)
                alanlar=[]
                radiusus=[]
                print(a)

                if circles is not None:
                    circles = np.uint16(np.around(circles))
                    for i in circles[0, :]:
                        center = (i[0], i[1])
                        # circle center
                        cv.circle(src, center, 1, (0, 100, 100), 3)
                        # circle outline
                        radius = i[2]
                        radiusus.append(radius*coefficientval)
                        cv.circle(src, center, radius, (255, 0, 255), 3)

                        def area():
                            global alan
                            alan= math.pi * ((radius*coefficientval) ** 2)
                        area()
                        alanlar.append(alan)

                measurebar()
                cv.destroyAllWindows()
            except Exception as e:
                print(e)
                popupmsg(e)
        sayaç=0
        def siyah_beyaz():
            src_color =(125, 125, 125)
            src_color1=(255,255,255)
            src_color2 =(125, 125, 125)
            src_color3=(0,0,0)
            image = cv.imread(basename1)
            # Compute matching pixels
            mask = cv.inRange(image, src_color, src_color1)
            image[mask > 0] = [255, 255, 255]
            mask = cv.inRange(image,  src_color3,src_color2)
            image[mask > 0] = [0, 0, 0]
            cv.imshow("s",image)
            cv.imwrite("kk.png",image)
        def browseFiles():
            global filename,basename1,vf_iter_basename1,plain_name,SoS_save_file_dir,SoS_save_file_results_of,SoS_save_file_results_filtered,SoS_save_file_results_result_of,SoS_save_file_results_excel,SoS_save_file_results_original
            filename = filedialog.askopenfilename()
            filepath = '{}'.format(filename)
            basename1 = os.path.basename(filepath)
            plain_name= os.path.splitext(basename1)[0]
            
            plain_name=plain_name+".xlsx"
            vf_iter_basename=os.path.basename(filepath)
            SoS_save_file_dir=str(os.path.join(os.environ["USERPROFILE"],"Desktop","VIQUAC","Save_File")).replace(os.sep, '/')
            SoS_save_file_results_of=str(os.path.join(os.environ["USERPROFILE"],"Desktop","VIQUAC","Save_File","Results_of_{}".format(basename1)))
            SoS_save_file_results_filtered=str(os.path.join(os.environ["USERPROFILE"],"Desktop","VIQUAC","Save_File","Results_of_{}".format(basename1),"Filtered_{}".format(basename1)))
            SoS_save_file_results_original=str(os.path.join(os.environ["USERPROFILE"],"Desktop","VIQUAC","Save_File","Results_of_{}".format(basename1),"Original_Image_of_{}".format(basename1)))
            SoS_save_file_results_result_of=str(os.path.join(os.environ["USERPROFILE"],"Desktop","VIQUAC","Save_File","Results_of_{}".format(basename1),"Result_of_{}".format(basename1)))
            SoS_save_file_results_excel=str(os.path.join(os.environ["USERPROFILE"],"Desktop","VIQUAC","Save_File","Results_of_{}".format(basename1),"{}".format(plain_name)))
        button_explore =ttk.Button(self,
                        text = "Browse Files",
                        command = browseFiles)
        button_explore.pack()

#opencv ascii karakterleri dışında karakterleri okumayıp ascii karakterli isimleri
#sıkıntısız okuyor onun dışındıaki dosyalarda ascii doğrulaması gerektiğini belirttir
#dosya isimleri dahil
        def direc():
            global dir_img
            try:
                global filename
                global refPt,buck_cast_list,ilk_sol_tık_y_list,ilk_sol_tık_x_list,istenen_noktalar_i_tipinden_2
                global cropping,img_for_sol_tık,bucket_image,bucket_demo_count,mode_7_event,tık_cast_dict
                global image,shifting,bucket_count,All_results,bucket_icin_scale_yapılmıs_mı,ilk_listeleri_icin_count
                global drawing,mode,start_x,start_y,end_x, end_y,r,img,r,bucket_list,dict_for_demo_buck_count,son_sol_tık_x_list,son_sol_tık_y_list
                global startAngle,endAngle,axesLength,center_coordinates,angle,shiftlenmis_mi,brush_olmayacak,istenen_noktalar_i_tipinden
                global scale_ratio,scale_count,img1,copy,mod_counter_for_other_than_5,reset_yemis_mi,bucketlan_mısmı
                global mode,counter_for_Scaled_brush,img_for_scaled_brush,brush_copy_first_counter,ii,selami,brush_cizilmismi,aaa,bucket_count_for_demo,istenen_noktalar_i_tipinden_for_demo,reset_for_buck
                shifting=False
                drawing = False#Drawing'in durumu True olması çizim yapıldığını bildirir.
                mode = 0 #Mode çizim araçları arasında geçiş yapabilmemiz için ufak bir parametre.
                x,y=-1000,-1000#Başlangıçta ileride çizilicek araçların window da gözükmemesi için ufak bir koordinat.
                start_x,start_y = -1,-1#Başlangıçta ileride çizilicek araçların window da gözükmemesi için ufak bir koordinat.
                end_x,end_y = 0,0#Başlangıçta ileride çizilicek araçların window da gözükmemesi için ufak bir koordinat.
                r = 3#Çemberin Yarıçapı için.
                angle=0#Ellipse'in açısı için.
                scale_ratio=100#Scaler fonksiyonunda ana parametre.
                scale_count=0#Scale içinde "+", "-" durumlarını anlasın diye.
                reset_for_buck=0
                shiftlenmis_mi=False
                brush_cizilmismi=False
                reset_yemis_mi=False
                mod_counter_for_other_than_5=2#Brush modunda Başlangıçta oluşabilecek hataları önlemek için *****EMİNLİK İÇERMEZ******.
                counter_for_Scaled_brush=False#Brush modunda scale işlemi yapılırken ekstra scalein o modda olduğunu bilsin diye
                #sanırım ********************* DÜZELTME GEREKLİ*********************
                brush_copy_first_counter=False
                ii=0
                brush_olmayacak=False
                bucketlan_mısmı=False
                istenen_noktalar_i_tipinden=[]
                bucket_count=0
                bucket_list=[]
                All_results=[]
                bucket_demo_count=0
                bucket_count_for_demo=0
                istenen_noktalar_i_tipinden_for_demo=[]
                mode_7_event=0
                dict_for_demo_buck_count={}
                buck_cast_list=[]
                bucket_icin_scale_yapılmıs_mı=False
                ilk_sol_tık_y_list=[]
                ilk_sol_tık_x_list=[]
                ilk_listeleri_icin_count=0
                istenen_noktalar_i_tipinden_2=[]
                
                def scaler():#******************************** Kalite sorunu SIKINTILI*************************
                    global scale_ratio,scale_count,img,copy,resized,img1,copy,ilk_sol_tık_y_list,ilk_sol_tık_x_list
                    global mod_counter_for_other_than_5,mode,img2,buck_cast_list,bucket_image,brush_cizilmismi,son_sol_tık_x_list,son_sol_tık_y_list
                    global img_for_scaled_brush,counter_for_Scaled_brush,brush_copy_first_counter,ii,selami,aaa,bucket_icin_scale_yapılmıs_mı

                    if scale_count==1:#Scale'in "-" olduğu kısımdır.
                        if scale_ratio==100:#Scale'in en son oran olarak 25'te kalmasını sağladık. Yoksa yazılım bozuluyor.
                            scale_ratio==100
                        else:
                            scale_ratio= scale_ratio - 100
                    elif scale_count==0:#Scale'in "+" olduğu kısımdır.
                        scale_ratio= scale_ratio + 100
                        
                    if brush_cizilmismi==True:
                        base_dim_img=np.copy(temp_img)
                        width = int(base_dim_img.shape[1] * scale_ratio / 100)
                        height = int(base_dim_img.shape[0] * scale_ratio / 100)
                        dim = (width, height)#Copy'nin boyutları alınıp aşağıdaki boyutlandırma fonksiyonuna giriyor.
                        img=cv.resize(aaa, dim, interpolation = cv.INTER_AREA)#Area interpolasyonu yapılıyor en boy oranı korunuyor.
                        img1=np.copy(img)#Img1 e img kopyalanıyor ne sebepse *********DÜZELTME GEREKEBİLİR************.
                        cv.imshow("image", img)#Resizelanan img window altında burda sergileniyor.

                    elif brush_cizilmismi==False:
                        img = np.copy(temp_img)#Img yi temp_img yapar.Template image hiç bir zaman içeriğini değiştirmiyorum.
                        copy = img.copy()#Copy burda template image oluyor.
                        width = int(copy.shape[1] * scale_ratio / 100)
                        height = int(copy.shape[0] * scale_ratio / 100)
                        dim = (width, height)#Copy'nin boyutları alınıp aşağıdaki boyutlandırma fonksiyonuna giriyor.
                        img=cv.resize(copy, dim, interpolation = cv.INTER_AREA)#Area interpolasyonu yapılıyor en boy oranı korunuyor.
                        img1=np.copy(img)#Img1 e img kopyalanıyor ne sebepse *********DÜZELTME GEREKEBİLİR************.
                        cv.imshow("image", img)#Resizelanan img window altında burda sergileniyor.                    
                    son_sol_tık_x_list=ilk_sol_tık_x_list[:]
                    son_sol_tık_y_list=ilk_sol_tık_y_list[:]
                    if bucketlan_mısmı==True:
                        bucket_icin_scale_yapılmıs_mı=True                   
                        for i in range(len(ilk_sol_tık_x_list)):
                            son_sol_tık_x_list[i]=int((scale_ratio/100)*ilk_sol_tık_x_list[i])
                            son_sol_tık_y_list[i]=int((scale_ratio/100)*ilk_sol_tık_y_list[i])
                        #print(son_sol_tık_x_list,son_sol_tık_y_list)
                        bucket_image=img.copy()  
                        bucket_herz()
                   
                def draw_shape(event,x,y,flags,param):
                    global drawing,mode,start_x,start_y,end_x, end_y,r,img,copy,Results,img_for_sol_tık,mode_7_event
                    global filename,startAngle,endAngle,axesLength,center_coordinates,reset_yemis_mi,shifting,brush_olmayacak
                    global angle,img1,copy,mod_counter_for_other_than_5,counter_for_Scaled_brush,img_for_scaled_brush,aaa,brush_cizilmismi

                    center_coordinates = (120, 100)#Bu ellipse için merkez koordinatlardır.
                    axesLength = (100, 50)#Ellipse için aksis uzunluklarıdır
                    startAngle = 0#Start-End angle Ellipse'in ne kadarının çizileceği tamamı için 0 360 bırakıyoruz.
                    endAngle = 360#Start-End angle Ellipse'in ne kadarının çizileceği tamamı için 0 360 bırakıyoruz.

                    if event == cv.EVENT_LBUTTONDOWN:#Mouse Bastığın an için.
                        #Bu Mouse'a basıldığında img yi Base Template kaydeder ki ileride kullanılırsa diye.

                        if mode==5:
                            pass
                        elif mode!=5:
                            if brush_cizilmismi==True:
                                pass
                            elif brush_cizilmismi==False:
                                if mode==7:
                                    pass
                                else:
                                    img = np.copy(img1)
                                    copy = img.copy()#Copy'ye img1 templateini kayıt eder her seferinde sol tık basıldığında resimi-
                        #yenileyip üst üste binmeleri önler.

                        drawing = True#Resim üstünde çizimin onayını paramtreye gömer.
                        start_x,start_y = x,y#Parametreye değer depolar.
                        end_x,end_y = x,y#Parametreye değer depolar.
                        if mode==7:
                            img_for_sol_tık=img.copy()
                            bucket_demo()
                            mode_7_event=1
                         
                    elif event == cv.EVENT_MOUSEMOVE and drawing:#Mouse Basılıyken hareketi algılar ve çizim le birlikte logic gate yapar.

                        end_x,end_y = x,y#Parametreye değer depolar.

                        if mode==1:#Line çizen branch-1.
                            cv.line(copy,(start_x,start_y),(end_x,end_y),(0,255,0),1)
                            cv.imshow('image', copy)

                        elif mode==2:#Rectangle çizen branch-2.
                            cv.rectangle(copy,(start_x,start_y),(end_x,end_y),(0,255,0),1)
                            cv.imshow("image", copy)

                        elif mode==3:#Circle çizen branch-3.
                            cv.circle(copy,(x,y),r,(0,0,255),1)
                            cv.imshow('image', copy)

                        elif mode==4:#Ellipse çizen branch-4.************GELİŞMESİ GEREK*************
                            if (end_y-start_y==0) and (end_x-start_x!=0):#Burda Ellipse için açısal değerlerin hata vermemesi için branch.
                                if end_x>start_x:
                                    angle=0
                                else :
                                    angle=0
                            elif (end_y-start_y!=0) and (end_x-start_x==0):
                                if end_y>start_y:
                                    angle=0
                                else:
                                    angle=0
                            else:
                                arc_value=(end_y-start_y)/(end_x-start_x)
                                angle =np.rad2deg(np.arctan(arc_value))#Ellipse için açı değeri hesaplar.
                            #a=(end_x-start_x)
                            #b=a/2
                            #axesLength1=(float(a),float(b))
                            #if a<0:
                                #a=(-1*a)
                                #b=(a/2)
                               # axesLength1=(float(a),float(b))
                            axesLength = (100, 50)
                            cv.ellipse(copy, (start_x,start_y), axesLength,angle,
                                0,360, (0,255,0), 1)
                            cv.imshow('image', copy)
                        #**********************************************************************
                        elif mode==5:#Brush yapan branch-5.
                                if brush_olmayacak==True:
                                    pass
                                else:
                                    reset_yemis_mi=False
                                    aaa=cv.circle(img,(x,y),r,(0,0,0),-1)#İçerideki parametreyi boyar.
                                    cv.imshow('image',aaa)#Scale de içerideki parametreyi windowa iter.
                        #**********************************************************************
                        elif mode==7:
                            cv.imshow('image',img_for_sol_tık)

                    elif event == cv.EVENT_LBUTTONUP:#Mouse tıklanmasının kalktığı harektlerin tamamlandığını belirtir.
                        drawing = False#Çizimin bittiğini drawinge depolar.
                        if mode==7:
                            mode_7_event=0
                            bucket_herz()                  
                def shifter():#Bela
                    global img,aaa,shiftlenmis_mi,scale_ratio
                    if shiftlenmis_mi==True:
                        pass
                    elif shiftlenmis_mi==False:
                        dim_background=img.copy()
                        width1 = int(dim_background.shape[1])
                        height1 = int(dim_background.shape[0])
                        width=width1*3
                        height=height1*3
                        gray_background=np.zeros([height,width,3],dtype=np.uint8)
                        gray_background.fill(255)
                        gray_background[height1:(2*height1),width1:(2*width1)]=img
                        img=gray_background.copy()
                        cv.imshow("image",img)
                        shiftlenmis_mi=True
                def radi_enlarger():#Yarıçap boyutlandırıcısı.
                    global r
                    if mode==2:#Çemberi büyütür.
                        r += 1
                    elif mode==2:#Çemberi küçültür.
                        if r <=2:
                            r = 1
                        else:
                            r -= 1
                def bucket_demo():
                    global dict_for_demo_buck_count,bucket_count_for_demo,bucket_image,bucket_demo_count,img_for_sol_tık,bucket_list,All_results,bucket_count,bucketlan_mısmı,img,start_x,start_y,brush_olmayacak,Results,istenen_noktalar_i_tipinden,istenen_noktalar_i_tipinden_for_demo
                    if bucket_demo_count==0:
                        bucket_image=img.copy()
                        bucket_demo_count= bucket_demo_count+1
                    imgray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
                    ret, thresh = cv.threshold(imgray, 127, 255, 0)
                    font = cv.FONT_HERSHEY_SIMPLEX
                    fontScale = 1
                    color = (0, 0, 0)
                    thickness = 1
                    contours,hierarchy = cv.findContours(thresh, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
                    for i in range(len(contours)):
                        cnt = contours[i]
                        answer=int(cv.pointPolygonTest(cnt,(start_x,start_y), False))
                        if answer==1:
                            M = cv.moments(contours[i])
                            cx = int(M['m10']/M['m00'])
                            cy = int(M['m01']/M['m00'])
                            if i in istenen_noktalar_i_tipinden_for_demo:
                                text_press=dict_for_demo_buck_count[i]
                                cv.putText(img_for_sol_tık,"{}".format(text_press),(cx,cy),font,fontScale, color, thickness, cv.LINE_AA)   
                            else:
                                istenen_noktalar_i_tipinden_for_demo.append(i)
                                bucket_count_for_demo=bucket_count_for_demo+1 
                                dict_for_demo_buck_count[i]=bucket_count_for_demo
                                cv.putText(img_for_sol_tık,"{}".format(bucket_count_for_demo),(cx,cy),font,fontScale, color, thickness, cv.LINE_AA)   
                    seed_point=(start_x,start_y)
                    new_val=(0,0,255)
                    cv.floodFill(img_for_sol_tık,None, seed_point, new_val)
                    cv.imshow('image',img_for_sol_tık)
                
                def bucket_herz():#**************************** BUCKET TAN SONRA + - gelirse sonra YAPILSINN***************
                    global scale_ratio,ilk_listeleri_icin_count,ilk_sol_tık_x_list,ilk_sol_tık_y_list,bucket_icin_scale_yapılmıs_mı,buck_cast_list
                    global bucket_demo_count,bucket_image,img_for_sol_tık,bucket_list,All_results,bucket_count,bucketlan_mısmı
                    global img,start_x,start_y,brush_olmayacak,Results,istenen_noktalar_i_tipinden,son_sol_tık_x_list,son_sol_tık_y_list,istenen_noktalar_i_tipinden_2
                    bucketlan_mısmı=True
                    seed_point=(start_x,start_y)
                    new_val=(0,0,255)

                    imgray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
                    ret, thresh = cv.threshold(imgray, 127, 255, 0)
                    contours,hierarchy = cv.findContours(thresh, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)                        
                    m_Centers=[]
                    font = cv.FONT_HERSHEY_SIMPLEX
                    fontScale = 1
                    color = (0, 0, 0)
                    thickness = 1
                    
                    
                    if bucket_icin_scale_yapılmıs_mı==True:
                        print(len(son_sol_tık_x_list))
                        bucket_count=0
                        brush_olmayacak=True
                        for i in range(len(son_sol_tık_x_list)):
                            seed_point=(son_sol_tık_x_list[i],son_sol_tık_y_list[i])
                            
                            for y in range(len(contours)):
                                cnt2 = contours[y]
                                answer=int(cv.pointPolygonTest(cnt2, (son_sol_tık_x_list[i],son_sol_tık_y_list[i]), False))#**********************Hata Burda
                                print(answer)
                                if answer==1:
                                    if y in istenen_noktalar_i_tipinden_2:#Burası Hatalı
                                        break
                                    else:
                                        M = cv.moments(cnt2)
                                        bucket_count= bucket_count+1
                                        if M['m00']==0 or M['m10']==0:
                                            pass
                                        else:
                                            istenen_noktalar_i_tipinden_2.append(y)
                                            cx = int(M['m10']/M['m00'])
                                            cy = int(M['m01']/M['m00'])
                                            cv.floodFill(bucket_image,None, seed_point, new_val)
                                            cv.putText(bucket_image,"{}".format(bucket_count),(cx,cy),font,fontScale, color, thickness, cv.LINE_AA)
                                            cv.imshow("image",bucket_image)
                                            print(seed_point,bucket_count)
                        bucket_icin_scale_yapılmıs_mı==False#****************************************************************************************************************
                    else:
                        if imgray[start_y,start_x]==0:
                            pass
                        else:
                            brush_olmayacak=True
                            for i in range(len(contours)):
                                cnt = contours[i]
                                answer=int(cv.pointPolygonTest(cnt, (start_x,start_y), False))
                                if answer==1:
                                    if i in istenen_noktalar_i_tipinden:
                                        break
                                    else:
                                        #cv.drawContours(bucket_image, contours, -1, (0,255,0), 1)
                                        istenen_noktalar_i_tipinden.append(i)
                                        bucket_count= bucket_count+1
                                        cnt1 = contours[i]
                                        M = cv.moments(cnt1)
                                        if M['m00']==0 or M['m10']==0:
                                            pass
                                        else:
                                            if i in istenen_noktalar_i_tipinden:
                                                cx = int(M['m10']/M['m00'])
                                                cy = int(M['m01']/M['m00'])
                                                m_Centers.append([cx,cy])
                                                cv.putText(bucket_image,"{}".format(bucket_count),(cx,cy),font,
                                                           fontScale, color, thickness, cv.LINE_AA)
                                                cv.floodFill(bucket_image,None, seed_point, new_val)
                                                x,y,w,h = cv.boundingRect(cnt1)
                                                area = cv.contourArea(cnt1)
                                                rect_area = w*h
                                                if area!=0 and rect_area!=0 :
                                                    extent = float(area)/rect_area
                                                    aspect_ratio = float(w)/h
                                                    hull = cv.convexHull(cnt1)
                                                    hull_area = cv.contourArea(hull)
                                                if hull_area!=0 or area!=0:
                                                    solidity = float(area)/hull_area
                                                    equi_diameter = np.sqrt(4*area/np.pi)
                                                cv.imshow("image",bucket_image)
                                                        
                                                Results=["Object {}:".format(bucket_count),"Area:{}".format(area),"Bounding Rectangle Area:{}".format(rect_area)
                                                ,"Extent:{}".format(extent),"Convex Hull Area:{}".format(hull_area),"Solidity:{}".format(solidity),"Equivalent Diameter:{}".format(equi_diameter)]
                                                All_results.append(Results)
                                                fake_scale_ratio=scale_ratio
                                                fake_scale_degismeyen_ratio=scale_ratio
                                                #print(start_x,fake_scale_degismeyen_ratio)
                                                if ilk_listeleri_icin_count==0 and bucket_icin_scale_yapılmıs_mı==False:
                                                    while True:
                                                        if fake_scale_ratio==100:
                                                            break
                                                        else :
                                                            fake_scale_ratio=fake_scale_ratio-100#zincir kuralına birdaha bak
                                                            base_x_val=start_x*(fake_scale_ratio/fake_scale_degismeyen_ratio)
                                                            base_y_val=start_y*(fake_scale_ratio/fake_scale_degismeyen_ratio)
                                                            #print("------",base_x_val,base_y_val)
                                                            continue
                                                    ilk_sol_tık_x_list.append(int(base_x_val))
                                                    ilk_sol_tık_y_list.append(int(base_y_val))
                                                    #print(ilk_sol_tık_x_list,ilk_sol_tık_y_list)
                                                    buck_cast_list.append(bucket_count) 
                                                
                def reset():#**********Her Reset de bucket listeleri de resetlenecek
                    global ilk_listeleri_icin_count,buck_cast_list,ilk_sol_tık_y_list,bucket_icin_scale_yapılmıs_mı
                    global ilk_sol_tık_x_list,ilk_listeleri_icin_count,dict_for_demo_buck_count
                    global All_results,bucket_count_for_demo,reset_yemis_mi,img,scale_count,brush_cizilmismi,istenen_noktalar_i_tipinden_2
                    global bucketlan_mısmı,brush_olmayacak,bucket_count,reset_for_buck,bucket_demo_count,istenen_noktalar_i_tipinden,istenen_noktalar_i_tipinden_for_demo
                    
                    bucket_count=0
                    bucket_demo_count=0
                    bucket_count_for_demo=0
                    istenen_noktalar_i_tipinden=[]
                    All_results=[]
                    istenen_noktalar_i_tipinden_for_demo=[]
                    reset_for_buck=1
                    dict_for_demo_buck_count={}
                    ilk_sol_tık_x_list=[]
                    ilk_sol_tık_y_list=[]
                    buck_cast_list=[]
                    ilk_listeleri_icin_count=0
                    bucket_icin_scale_yapılmıs_mı=False
                    istenen_noktalar_i_tipinden_2=[]
                   
                    if brush_cizilmismi==True:#bucket bunun içinde olacak
                        if bucketlan_mısmı==True:
                            brush_olmayacak=False
                        brush_cizilmismi=False
                        bucketlan_mısmı=False
                        reset_yemis_mi=True
                        scale_count=3
                        scaler()
                        #1 resetlenmiş 0 resetlenmeye hazır demek
                        cv.imshow('image',img)
                    else:
                        if bucketlan_mısmı==True:
                            scale_count=3
                            scaler()
                            cv.imshow('image',img)
                            bucketlan_mısmı=False
                            brush_olmayacak=False
                            reset_yemis_mi=True
                            
                        cv.imshow('image',img)
                def cut_cal(img,x,y,x_,y_):
                    try:
                        global filename 
                        
                        title=filename.replace(os.sep, '/')
                        title=title.split("/")
                        title_Name=title[-1]
                        dir_list=title.pop(-1)
                        dir_="/".join(title)
                        print(dir_)
                        #foto sağdan sola croplanmıyor onu çöz ve yukarı doğru olmuyor
                        org_img_x,org_img_y,z=img.shape
                        cropped_img=img[y:y_,x:x_]
                        x_val,y_val,z=cropped_img.shape
                        
                        val={}
                        
                        for i in range(y_val):
                            val[i]=[]
                            for c in range(x_val):
                                # print(c)
                                s=np.mean(cropped_img[c,i])
                                val[i].append(s)
                        new_val={}
                        y=[]
                        x=[]
                        for i in range(y_val):
                            new_val[i]=np.mean(val[i])
                            y.append(new_val[i])
                            x.append(i)
                       
                        fig,ax=plt.subplots()
                        plt.ylabel('Gray Scale Values')
                        plt.xlabel('Distance (x-Axis)')
                        plt.title("{}".format(title_Name))
                        ax.plot(x,y)
                        plt.savefig("{}/{}_Gray_Scale_Value_Graph{}".format(dir_,title_Name,".png"))
                        plt.show()
                        
                        # return cropped_img
                    except Exception as E:
                        print(E)

                img=cv.cvtColor(dir_img, cv.COLOR_GRAY2RGB)#Başta yüklenen fotoğraf.
                img1=np.copy(img)#Başta yüklenen fotoğrafın kopyası burdan kod 1 kere geçiyor.
                temp_img = np.copy(img)#Başta yüklenen fotoğrafın kopyası burdan kod 1 kere geçiyor.
                cv.namedWindow('image')#Ana while penceresi.
                cv.setMouseCallback('image',draw_shape)#Pencere açıkken içinde mouse için Callback.
                
                while(1):#Döndür abim dönsün.
                    if not cv.EVENT_MOUSEMOVE:#Mouse Hareketi yoksa demek.
                        if mode == 0:#Eğer Line modundaysa.
                            copy = img.copy()
                            cv.imshow('image', copy)
                        elif mode==1:
                            copy = img.copy()
                            cv.line(copy,(start_x,start_y),(end_x,end_y),(0,255,0),1)
                            cv.imshow('image', copy)
                        elif mode==2:#Eğer Rectangle modundaysa.
                            copy = img.copy()
                            cv.rectangle(copy,(start_x,start_y),(end_x,end_y),(0,255,0),1)
                            cv.imshow('image',copy)
                        elif mode==3:#Eğer Circle modundaysa.
                            copy = img.copy()
                            cv.circle(copy,(end_x,end_y),r,(0,0,255),1)
                            cv.imshow('image',copy)
                        elif mode==4:#Eğer Ellipse modundaysa.
                            copy = img.copy()
                            cv.ellipse(copy, (start_x,start_y), axesLength,angle,
                                    0,360, (0,255,0), 1)
                            cv.imshow('image',copy)
                        elif mode==5:#Eğer Brush modundaysa.
                            if brush_cizilmismi==False:
                                if brush_olmayacak==True:#mode 7 için
                                    pass
                                else:
                                    if reset_yemis_mi==True:
                                        cv.imshow('image',img)
                                    else:#buraya başta brush count eklenip başlangıç olu olmadığı anlaşılabilir
                                        copy = img.copy()
                                        aaa=cv.circle(copy,(x,y),r,(0,0,0),-1)#İçerideki parametreyi boyar.'''
                                        cv.imshow('image',img)#Scale de içerideki parametreyi windowa iter.
                                        brush_cizilmismi=True
                            elif brush_cizilmismi==True:
                                if reset_yemis_mi==True:
                                    copy = img.copy()
                                    cv.imshow('image',copy)
                                    reset_yemis_mi=False
                                elif reset_yemis_mi==False:
                                    aaa=cv.circle(img,(x,y),r,(0,0,0),-1)#İçerideki parametreyi boyar.
                                    cv.imshow('image',aaa)#Scale de içerideki parametreyi windowa iter.
                                    #**********************************************************************
                        elif mode==6:#Shifting
                            cv.imshow("image",img)
                        elif mode==7:
                            #buraya resetlendiğinde gösterilecek imaj için yer gelecek
                            if mode_7_event==1:
                                if bucket_demo_count!=0:
                                    cv.imshow("image",img_for_sol_tık)
                            elif mode_7_event==0:
                                if bucket_demo_count==0:
                                    cv.imshow("image",img)
                                elif bucket_demo_count!=0:
                                    cv.imshow("image",bucket_image),
                                
                                    
                    k = cv.waitKeyEx(1) & 0xFF #Pencere içindeki anahtar girişi için gereken bir fonksiyon.

                    if k == ord('w'): #Q Line modu için.
                        mode=1
                        end_x,end_y = -10,-10
                        start_x,start_y = -10,-10
                    elif k == ord('q'):#W Rectangle modu için.
                        mode=2
                        end_x,end_y = -10,-10
                        start_x,start_y = -10,-10
                    elif k==ord("a"):
                        cv.destroyAllWindows()
                        cut_cal(img,start_x,start_y,end_x,end_y)
                        break
                        
                    elif k == ord('e'):#E Circle modu için.
                        mode=3
                        end_x,end_y = -10,-10
                        start_x,start_y = -10,-10
                    elif k ==ord("b"):#B Brush  modu için.
                        mode=5
                        end_x,end_y = -10,-10
                        start_x,start_y = -10,-10
                    elif k == ord('r'):#Ellipse modu için.
                        mode=4
                        end_x,end_y = -10,-10
                        start_x,start_y = -10,-10
                    elif k == ord('s') and (mode==3 or mode==5):#S Circle Büyütücü için :).
                        r += 1
                    elif k == ord("f") and (mode==3 or mode==5):#F Circle Küçültücü için :(.
                        if r <=2:
                            r = 1
                        else:
                            r -= 1
                    elif k == ord('x'):#X Mode 5 hariç resetlemek için DÜZELTİLECEK.**********************
                        end_x,end_y = -10,-10
                        start_x,start_y = -10,-10
                        reset()
                    elif k == ord('+'):#"+" Zoom-In için.
                        scale_count=0
                        scaler()
                        #Bunu Bir araştır*************************
                    elif k == ord('-'):#"-" Zoom-Out için
                        scale_count=1
                        scaler()
                    elif k==ord("p"):
                        shifter()
                    elif k==ord("m"):
                        mode=7
                    elif k==ord("h"):
                        for i in range(len(All_results)):
                            Reuslt_Window=tk.Toplevel()
                            Reuslt_Window.geometry('300x300')
                            Reuslt_Window.title("VIQUAC")
                            for z in range(len(Results)):
                                ttk.Label(Reuslt_Window,text=All_results[i][z]).pack(anchor="w")
                    elif k == 27:#"27" Escapein Karşılığı için.
                        break
                    
                cv.destroyAllWindows()
                #cv.imshow('image',bucket_image)# bu sadece sonda alan hesabında göstermek için sonrasında fotoğraf yanı gibi buna da fotoğraf altı konulabilinir.
                img=cv.cvtColor(dir_img, cv.COLOR_GRAY2RGB)

            except Exception as e:
                print(e)
                #popupmsg("Please Select an Image")
        def directionality():
            global filename
            global basename1
            global popupmsg
            global dir_img

            try:
                global matrice_order

                dir_img=cv.imread(filename,0)#gray scale de tek

                dir_wid = dir_img.shape[1]
                dir_hgt = dir_img.shape[0]

                dir_col_num=[]#renkleri paketlediğimiz liste

                for y in range (0,dir_hgt,1):
                    for x in range(0,dir_wid,1):
                        dir_col_num.append(dir_img[y,x])

                DN=1#ortaya gelenlerin numaraları işlem sıralarına göre tabi ki

                wid_border=dir_wid-2

                hgt_border=dir_hgt-2

                matrice_order=[]

                for y in range(hgt_border):
                    for x in range(wid_border):
                        a1=(y*(dir_wid))+x
                        a2=(y*(dir_wid))+(x+1)
                        a3=(y*(dir_wid))+(x+2)
                        b1=((y+1)*(dir_wid))+x
                        b2=((y+1)*(dir_wid))+(x+1)
                        b3=((y+1)*(dir_wid))+(x+2)
                        c1=((y+2)*(dir_wid))+x
                        c2=((y+2)*(dir_wid))+(x+1)
                        c3=((y+2)*(dir_wid))+(x+2)
                        matrice_order.append([a1,a2,a3,b1,b2,b3,c1,c2,c3])

                matrice_order=np.array(matrice_order)

                '''
                d1f=pd.DataFrame(matrice_order)
                filep="matrice_order.xlsx"
                d1f.to_excel(filep,index=False)'''



                def line_to_matrice_converter(i):#burda değerler hep +1 alarak okunacak kafadan itere sıkıntı yapıyor ondan sıkıntı etmeyyin

                    global matrice_order
                    global DN_X
                    global DN_Y
                    global remain
                    global bölüm


                    #print("değer:{}".format(matrice_order[i][1]))
                    #print("wid-1:{}".format(dir_wid-1))

                    remain=int(((matrice_order[i][4]))%(dir_wid))

                    bölüm=int(((matrice_order[i][4]))/(dir_wid))
                    if remain!=0 and bölüm!=0:
                        DN_X=bölüm
                        DN_Y=remain
                        #print(DN_X,DN_Y,"1")
                    elif remain==0 and bölüm == 0:
                        DN_X=0
                        DN_Y=0
                        #print(DN_X,DN_Y,"2")
                    elif remain!=0 and bölüm==0:
                        DN_X=bölüm+1
                        DN_Y=bölüm+1
                        #print(DN_X,DN_Y,"3")
                    elif remain==0 and bölüm!=0:
                        DN_X=bölüm-1
                        DN_Y=dir_wid-2
                        #print(DN_X,DN_Y,"4")
                """            
                src_color =(0)
                src_color1=(100)
                src_color2=(100)
                src_color3=(255)

                mask = cv.inRange(dir_img, src_color, src_color1)
                mask1 = cv.inRange(dir_img, src_color2, src_color3)

                dir_img[mask > 0] = [0]
                dir_img[mask1 > 0] = [255]"""

                kernel = np.ones((5,5),np.float32)/25
                # dst = cv.filter2D(dir_img,-1,kernel)
                for i in range(len(matrice_order)):
                    #bu yer olduğundan esas dim de değer okumaya bakılacak
                    line_to_matrice_converter(i)

                    DN_MIDDLE=dir_img[DN_X,DN_Y]

                    DN_LEFT=dir_img[DN_X-1,DN_Y]

                    DN_RIGHT=dir_img[DN_X+1,DN_Y]

                    DN_DOWN=dir_img[DN_X,DN_Y-1]

                    DN_UP=dir_img[DN_X,DN_Y+1]

                    DN_UP_LEFT=dir_img[DN_X-1,DN_Y+1]

                    DN_UP_RIGHT=dir_img[DN_X+1,DN_Y+1]

                    DN_DOWN_LEFT=dir_img[DN_X-1,DN_Y-1]

                    DN_DOWN_RIGHT=dir_img[DN_X+1,DN_Y-1]

                    if DN_MIDDLE==255:#alt satırlar sadece tek bir siyah noktanın yalnız kaldığını silecek
                        if DN_LEFT==0 and DN_RIGHT==0 and DN_UP==0 and DN_DOWN==0 :
                            dir_img[DN_X,DN_Y]=0



                direc()


                '''
                Thick_Calculater=original = cv.imread(filename)
                wid = Thick_Calculater.shape[1]
                hgt = Thick_Calculater.shape[0]
                #print(wid,hgt)
                def sunny_Side_up(wid,hgt):
                    x=wid
                    y=hgt
                    def GCD(a,b):
                        if a == 0 :
                            return b
                        return GCD(b%a,a)
                    Total_Area_of_Image=x*y
                    GCD_Value=GCD(x,y)
                    Side_Lengths_of_The_Square=[]
                    Numbers_of_Reference_Square_List=[]
                    for i in range(GCD_Value,0,-1):
                        Numbers_of_Reference_Square=(Total_Area_of_Image/i**2)
                        if Numbers_of_Reference_Square.is_integer():
                            Side_Lengths_of_The_Square.append(i)
                            Numbers_of_Reference_Square_List.append(Numbers_of_Reference_Square)
                    for i in range(len(Side_Lengths_of_The_Square)):
                        print("Length:{}".format(Side_Lengths_of_The_Square[i]),"Number:{}".format(Numbers_of_Reference_Square_List[i]))


                sunny_Side_up(wid,hgt)
                cv.namedWindow("{}".format(basename1))
                while True:
                    if vf_iter_counter==0:
                        original = cv.imread(filename)
                    image = original.copy()
                    k = cv.waitKey(1) & 0xFF
                    if k == 27:
                        cv.destroyAllWindows()
                        break
                    cv.resizeWindow("{}".format(basename1),1920,1080)
                    cv.imshow("{}".format(basename1),image)'''
            except Exception as e:
                print(e)
                popupmsg("Two errors can be occur:\n    1-Please upload your image.\n    If it doesn't upload your image then check these steps:\n            1.1-Please use Ascii characters in your file name and directory.\n        1.2-Please check your file name integrity\n    2-Press 'Escape' button to close the window when your work done.")
        direct_button=ttk.Button(self,text="Directionality",command=directionality)
        direct_button.pack()
        def yes_or_no():
            global answer_of_yes_or_no
            yes_no=tk.Toplevel()
            yes_no.geometry('400x100')
            yes_no.title("VIQUAC")
            ttk.Label(yes_no,text="Vf found from previous process. Do you want to recalculate?").pack(expand=True)
            answer_of_yes_or_no=0
            def cavani_porto_Bene_yes():
                global answer_of_yes_or_no
                answer_of_yes_or_no=1
                yes_no.destroy()
            def cavani_porto_Bene_no():
                global answer_of_yes_or_no
                answer_of_yes_or_no=0
                yes_no.destroy()
            ttk.Button(yes_no,text='Yes',command=cavani_porto_Bene_yes).pack(expand=True)
            ttk.Button(yes_no,text='No',command=cavani_porto_Bene_no).pack(expand=True)
            yes_no.wait_window()
        def volfrac():
            global basename1
            global pixcont
            global vf_iter_counter
            global kızart
            global kırmızı
            global brightness
            global contrast
            global frac_sayac
            global filename
            global plain_name
            try:
                global Vm
                global Vf
                global effect
                global yes
                global no
                global answer_of_yes_or_no
                global hgt,wid
                global volfrac_calculator_counter
                if 'basename1' in globals() :
                    print(os.path.exists(SoS_save_file_dir))
                    if os.path.exists(SoS_save_file_dir):
                        if os.path.exists(SoS_save_file_results_excel):
                            excel_workbook=openpyxl.load_workbook(SoS_save_file_results_excel)
                            mec_feature_page=excel_workbook.active

                            Vf=mec_feature_page.cell(row=1,column=2).value
                            Vm=mec_feature_page.cell(row=2,column=2).value

                            if (1>Vf>0 or 1>Vm>0) and Vf+Vm==1:
                                yes_or_no()
                                if answer_of_yes_or_no==0:
                                    effect=cv.imread(SoS_save_file_results_filtered)
                                    wid = effect.shape[1]
                                    hgt = effect.shape[0]
                                    return Vf,Vm,effect,hgt,wid

                            else:
                                if volfrac_calculator_counter==1:
                                    volfrac_calculator_counter=0
                                    pass
                                else:
                                    popupmsg("Your volume fraction values has been changed by external factors. Please recalculate the values.\n Volume Fraction of Fibers: {}\n Volume Fraction of Fibers: {}".format(Vf,Vm))
                                    pass
                    else:
                        os.mkdir("C:/Users/deniz/Desktop/SoS/Save_File")
                    def pixcont(image, kırmızı=255,brightness=550,contrast=127):#def pixcont(image, kırmızı=255,brightness=550,contrast=127):

                        global basename1,filename,filenam,x111,ss,cc,hgt,wid,outputImage

                        brightness = int((brightness - 0) * (255 - (-255)) / (510 - 0) + (-255))
                        contrast = int((contrast - 0) * (127 - (-127)) / (254 - 0) + (-127))
                        if brightness != 0:
                            if brightness > 0:
                                shadow = brightness
                                max = 255
                            else:
                                shadow = 0
                                max = 255 + brightness
                            al_pha = (max - shadow) / 255
                            ga_mma = shadow
                    		# The function addWeighted calculates
                    		# the weighted sum of two arrays
                            cal = cv.addWeighted(image, al_pha,
                    							image, 0, ga_mma)
                        else:
                            cal = image
                        if contrast != 0:
                            Alpha = float(131 * (contrast + 127)) / (127 * (131 - contrast))
                            Gamma = 127 * (1 - Alpha)
                    		# The function addWeighted calculates
                    		# the weighted sum of two arrays
                            cal = cv.addWeighted(cal, Alpha,
                    							cal, 0, Gamma)

                        wid = image.shape[1]
                        hgt = image.shape[0]

                        ss=[]
                        x1=[]
                        cc=[]
                        src_color =(kırmızı, kırmızı, kırmızı)
                        src_color1=(255,255,255)

                        image=cal

                        # Compute matching pixels
                        mask = cv.inRange(image, src_color, src_color1)

                        # Replace pixels where the mask is nonzero
                        if black==0:
                            if kırmızı==0:
                                image[mask > 0] = [255, 255, 255]
                            else:
                                image[mask > 0] = [0, 0, 255]
                        else:
                            if kırmızı==0:
                                image[mask > 0] = [255, 255, 255]
                            else:
                                image[mask < 255] = [0, 0, 255]

                        global kır
                        global brig
                        global cont
                        kır=kırmızı
                        brig=brightness
                        cont=contrast
                        return image
                    fotoolcer=0
                    def dead(x):
                        pass
                    cv.namedWindow('Filter')

                    cv.createTrackbar('Brightness','Filter', 255,2*255,dead)
                    cv.createTrackbar('Red Area Ratio','Filter', 127,255,dead)
                    cv.createTrackbar('Contrast','Filter',127, 2 * 127,dead)

                    while True:
                        if vf_iter_counter==1:
                            original = cv.imread(vf_iter_basename)

                        elif vf_iter_counter==0:
                            original = cv.imread(filename)

                        image = original.copy()

                        #image = cv.cvtColor(image, cv.COLOR_BGR2GRAY)

                        #cv.resizeWindow('Filter',hgt,wid)

                        k = cv.waitKey(1) & 0xFF
                        if k == 27:
                            break

                        kırmızı = cv.getTrackbarPos('Red Area Ratio','Filter')
                        brightness = cv.getTrackbarPos('Brightness','Filter')
                        contrast = cv.getTrackbarPos('Contrast','Filter')
                        effect = pixcont(image,kırmızı,brightness,contrast)
                        #cv.resize(effect,(1920,1080))
                        #cv.resizeWindow('Filter',1920,1080)
                        cv.imshow('Filter',effect)

                        if os.path.exists(SoS_save_file_results_of):
                            cv.imwrite(SoS_save_file_results_original,original)
                            cv.imwrite(SoS_save_file_results_filtered,effect)
                        else:
                            os.mkdir(SoS_save_file_results_of)
                            cv.imwrite(SoS_save_file_results_original,original)
                            cv.imwrite(SoS_save_file_results_filtered,effect)

                    cv.destroyAllWindows()

                    #burda alan hesabını butona çekki kasma yapmasun
                    Redones=0

                    for y in range(0,hgt,1):
                        for x in range (0,wid,1):
                            if (effect[y,x]==[0,0,255]).all():
                                Redones=Redones+1

                    alan=hgt*wid
                    Vf=(Redones/alan)
                    Vm=1-Vf
                    position = (wid+10,50)

                    #excel_workbook=openpyxl.load_workbook('C:/Users/deniz/OneDrive/Desktop/SoS/Save_File/{}'.format(plain_name))
                    excel_workbook=openpyxl.Workbook()
                    del excel_workbook['Sheet']
                    mec_feature_page=excel_workbook.create_sheet("Mechanical_Features",0)
                    #0 ilk sırada başlamasını sağlar istediğin değeri ver -1 sondan önceki modu
                    openpyxl.worksheet.dimensions.ColumnDimension("Mechanical_Features",index='A',width=float(100))
                    mec_feature_page.column_dimensions['A'].width = 30
                    mec_feature_page_title_volume_fraction_of_fiber=mec_feature_page.cell(row=1,column=1,value="Volume Fraction of Fiber")
                    mec_feature_page_value_volume_fraction_of_fiber=mec_feature_page.cell(row=1,column=2,value=Vf)
                    mec_feature_page_title_volume_fraction_of_resin=mec_feature_page.cell(row=2,column=1,value="Volume Fraction of Resin")
                    mec_feature_page_value_volume_fraction_of_resin=mec_feature_page.cell(row=2,column=2,value=Vm)
                    excel_workbook.save(filename = SoS_save_file_results_excel)

                elif 'basename1' not in globals():
                    print(4)
                    popupmsg("Please upload an image to run the code")
                    
                        
            except Exception as e:
                print(e)
                #popupmsg("Two errors can be occur:\n    1-Please upload your image.\n    If it doesn't upload your image then check these steps:\n            1.1-Please use Ascii characters in your file name and directory.\n        1.2-Please check your file name integrity\n    2-Press 'Escape' button to close the window when your work done.")
        def calculator():
            global pixcont
            global kır
            global brig
            global cont
            global kızart
            try:
                global outputImage
                global Vm
                global Vf
                global hgt
                global wid
                global pixcont
                global kırmızı
                global brightness
                global contrast
                global effect
                global hgt,wid
                global volfrac_calculator_counter
                volfrac_calculator_counter=0
                if 'basename1' in globals() :
                    if os.path.exists(SoS_save_file_dir):
                        # burda dosya ismini check edebiliriz
                        if os.path.exists(SoS_save_file_results_of):
                            if os.path.exists(SoS_save_file_results_excel):
                                excel_workbook=openpyxl.load_workbook(SoS_save_file_results_excel)
                                mec_feature_page=excel_workbook.active

                                Vf=mec_feature_page.cell(row=1,column=2).value
                                Vm=mec_feature_page.cell(row=2,column=2).value

                                if (1>Vf>0 or 1>Vm>0) and Vf+Vm==1:

                                    effect=cv.imread(SoS_save_file_results_filtered)
                                    wid = effect.shape[1]
                                    hgt = effect.shape[0]
                                else:
                                    popupmsg("Your volume fraction values has been changed by external factors. Please recalculate the values.\n Volume Fraction of Fibers: {}\n Volume Fraction of Fibers: {}".format(Vf,Vm))
                                    volfrac_calculator_counter=1
                                    volfrac()
                                    #sayaç eklenecek
                                    pass
                            else:
                                popupmsg("You don't have any saves as {} in your directory.Volume fraction calculations and mechanical behavior determinations will be started.".format(plain_name))
                                volfrac_calculator_counter=1
                                volfrac()
                                pass
                        else:
                            popupmsg("You don't have any saves as Results_of_{} in your Save_File directory.Volume fraction calculations and mechanical behavior determinations will be started.".format(basename1))
                            volfrac_calculator_counter=1
                            volfrac()
                            pass

                    else:
                        popupmsg("You don't have Save_File in your directory.The Save_file and Results_of_{} will be created.Volume fraction calculations and mechanical behavior determinations will be started.".format(basename1))
                        os.mkdir(SoS_save_file_dir)
                        volfrac_calculator_counter=1
                        volfrac()
                        pass

                        #ask for function eğer yapılcaksa volfrac istesin
                    if hgt<=450:
                        outputImage = cv.copyMakeBorder(
                         effect,
                         0,
                         200,
                         0,
                         1000,
                         cv.BORDER_CONSTANT,
                         value=(255,255,255)
                      )

                    elif hgt>450:
                        outputImage = cv.copyMakeBorder(
                         effect,
                         0,
                         0,
                         0,
                         1000,
                         cv.BORDER_CONSTANT,
                         value=(255,255,255)
                      )
                    global fibers
                    global resins
                    fibers=["AS 4","T 300","E-Glass 21 x K43 GEVETEX","Silenka E-Glass 1200 Tex"]
                    resins=["3501-6 epoxy","BSL914C epoxy","LY556/HT907/DY063 Epoxy ","MY750/HY917/DY063 Epoxy ","8552 Epoxy"]

                    def calculater_s_in():
                        global selectedfib
                        global selectedmat
                        global E11
                        global E22
                        global In_plane_V12
                        global In_plane_shear_modulus

                        data=sql.connect("Material_Library_new.sqlite")
                        fib=data.cursor()
                        fib.execute("""SELECT * FROM Fibers""")
                        mat=data.cursor()
                        mat.execute("""SELECT * FROM Resins""")
                        datafib=fib.fetchall()
                        datamat=mat.fetchall()

                        for row in datafib:
                            if selectedfib in row:
                                Ef1=row[1]
                                Ef2=row[2]
                                Gf12=row[3]
                                vf12=row[4]
                                Gf23=row[5]
                                Lon_Tens_Strength=row[6]
                                Lon_Comp_Strength=row[7]
                                Lon_Tens_Fail_Strain=row[8]
                                Lon_Comp_Fail_Strain=row[9]
                                alfa1=row[10]
                                alfa2=row[11]
                        for row2 in datamat:
                            if selectedmat in row2:
                                Em=row2[1]
                                Gm=row2[2]
                                vm=row2[3]
                                Ymt=row2[4]
                                Ymc=row2[5]
                                Sm=row2[6]
                                Emt=row2[7]
                                Ym=row2[8]
                                alfam=row2[9]

                        E11=Ef1*Vf+Em*Vm
                        E22=(Ef2*Em)/((Vf*Em)+(Vm*Ef2))
                        In_plane_V12=vf12*Vf+Vm*vm
                        In_plane_shear_modulus=Gf12*Gm/(vf12*Gm+vm*Gf12)

                        textim=["Assigned Fiber Type : {} ".format(selectedfib),"Assigned Resin Type : {} ".format(selectedmat),"Fiber Volume Fraction : {}%".format(Vf*100),"Resin Volume Fraction : {}%".format(Vm*100),"Longitudinal Modulus: {} [GPa]".format(E11),"Transverse modulus : {} [GPa]".format(E22),"In-plane Poisson ratio : {}".format(In_plane_V12),"In-plane shear modulus : {} [GPa]".format(In_plane_shear_modulus)]
                        #print(textim)
                        hgt=25
                        position = (wid+10,hgt)

                        pokus=outputImage.copy()
                        for i in range(len(textim)):
                            position = (wid+10,hgt)
                            cv.putText(
                                 pokus,
                                 textim[i],
                                 position,
                                 cv.FONT_HERSHEY_SIMPLEX,
                                 1,
                                 (0, 0, 0, 255),
                                 1)

                            hgt=hgt+50
                        data.commit()
                        global predict_count
                        if predict_count==1:
                            predict_count=0
                            pass
                        elif predict_count==0:
                            cv.imwrite(SoS_save_file_results_result_of, pokus)

                    def fiber_resin_picker():
                        global fibers
                        global resins
                        global fib_res_picker
                        global predict_count
                        global Vf
                        global Vm

                        predict_count=0
                        fib_res_picker=tk.Toplevel()
                        fib_res_picker.geometry('400x300')
                        fib_res_picker.title("VIQUAC")

                        fibers_list=[]
                        selectedfib_number=tk.IntVar()
                        resins_list=[]
                        selectedres_number=tk.IntVar()
                        print(fibers)
                        ttk.Label(fib_res_picker,text="Select Fiber Type").pack(anchor="w")
                        for i  in range(len(fibers)):
                            fibers_list.append(ttk.Checkbutton(fib_res_picker,text="{}".format(fibers[i]),onvalue=i,variable=selectedfib_number))
                            fibers_list[i].pack(anchor="w")

                        ttk.Label(fib_res_picker,text="Select Resin Type").pack(anchor="w")
                        for i  in range(len(resins)):
                            resins_list.append(ttk.Checkbutton(fib_res_picker,text="{}".format(resins[i]),onvalue=i,variable=selectedres_number))
                            resins_list[i].pack(anchor="w")

                        def cavani_porto_Bene_approve():
                            global selectedfib
                            global selectedmat
                            selectedfib=fibers[selectedfib_number.get()]
                            selectedmat=resins[selectedres_number.get()]
                            fib_res_picker.destroy()

                        def cavani_porto_Bene_predict():
                            global selectedfib
                            global selectedmat
                            selectedfib=fibers[selectedfib_number.get()]
                            selectedmat=resins[selectedres_number.get()]
                            calculater_s_in()
                            list_of_behavior=["Assigned Fiber Type:{}".format(selectedfib),"Assigned Fiber Type:{}".format(selectedmat),"Vf:{}".format(Vf),"Vm:{}".format(Vm),"Longitudinal Modulus:{}".format(E11),"Tenisle Modulus:{}".format(E22),"In-Plane Poisson's Ratio:{}".format(In_plane_V12),"In-Plane Shear Modulus:{}".format(In_plane_shear_modulus)]

                            predict_window=tk.Toplevel()
                            predict_window.geometry('300x300')
                            predict_window.title("VIQUAC")
                            for i in range(len(list_of_behavior)):
                                ttk.Label(predict_window,text=list_of_behavior[i]).pack(anchor="w")

                        ttk.Button(fib_res_picker,text='Predict',command=cavani_porto_Bene_predict).pack()
                        ttk.Button(fib_res_picker,text='Confirm',command=cavani_porto_Bene_approve).pack()
                    fiber_resin_picker()
                    fib_res_picker.wait_window()
                    calculater_s_in()
                    excel_workbook=openpyxl.load_workbook(SoS_save_file_results_excel)
                    mec_feature_page=excel_workbook.active
                    
                    #0 ilk sırada başlamasını sağlar istediğin değeri ver -1 sondan önceki modu

                    mec_feature_page.column_dimensions['A'].width = 30
                    mec_feature_page.column_dimensions['B'].width = 30


                    mec_feature_page_title_assigned_Fiber_type=mec_feature_page.cell(row=3,column=1,value="Assigned Fiber Type")
                    mec_feature_page_value_assigned_Resin_type=mec_feature_page.cell(row=3,column=2,value=selectedfib)

                    mec_feature_page_title_assigned_Resin_type=mec_feature_page.cell(row=4,column=1,value="Assigned Resin Type")
                    mec_feature_page_value_assigned_Resin_type=mec_feature_page.cell(row=4,column=2,value=selectedmat)

                    mec_feature_page_title_Longitudinal_Modulus=mec_feature_page.cell(row=5,column=1,value="Longitudinal Modulus (GPa)")
                    mec_feature_page_value_Longitudinal_Modulus=mec_feature_page.cell(row=5,column=2,value=E11)

                    mec_feature_page_title_Transverse_Modulus=mec_feature_page.cell(row=6,column=1,value="Transverse Modulus (GPa)")
                    mec_feature_page_value_Transverse_Modulus=mec_feature_page.cell(row=6,column=2,value=E22)

                    mec_feature_page_title_In_Plane_Poisson_Ratio=mec_feature_page.cell(row=7,column=1,value="In-Plane Poisson Ratio")
                    mec_feature_page_value_In_Plane_Poisson_Ratio=mec_feature_page.cell(row=7,column=2,value=In_plane_V12)

                    mec_feature_page_title_In_Plane_Shear_Modulus=mec_feature_page.cell(row=8,column=1,value="In-Plane Shear Modulus (GPa)")
                    mec_feature_page_value_In_Plane_Shear_Modulus=mec_feature_page.cell(row=8,column=2,value=In_plane_shear_modulus)


                    excel_workbook.save(filename = SoS_save_file_results_excel)
                elif 'basename1' not in globals():
                    popupmsg("Please upload an image to run the code")

            except Exception as e:
                popupmsg(e)
                fib.close()
                mat.close()
        def defecter():
            
            try:
                def pixcont(image, kırmızı=255,brightness=550,
                            			contrast=127):
                    global basename1
                    global ss
                    global x111
                    global cc
                    global hgt
                    global wid
                    black=0
                    brightness = int((brightness - 0) * (255 - (-255)) / (510 - 0) + (-255))
                    contrast = int((contrast - 0) * (127 - (-127)) / (254 - 0) + (-127))
                    if brightness != 0:
                        if brightness > 0:
                            shadow = brightness
                            max = 255
                        else:
                            shadow = 0
                            max = 255 + brightness
                        al_pha = (max - shadow) / 255
                        ga_mma = shadow
                		# The function addWeighted calculates
                		# the weighted sum of two arrays
                        cal = cv.addWeighted(image, al_pha,
                							image, 0, ga_mma)
                    else:
                        cal = image
                    if contrast != 0:
                        Alpha = float(131 * (contrast + 127)) / (127 * (131 - contrast))
                        Gamma = 127 * (1 - Alpha)
                		# The function addWeighted calculates
                		# the weighted sum of two arrays
                        cal = cv.addWeighted(cal, Alpha,
                							cal, 0, Gamma)


                    wid = image.shape[1]
                    hgt = image.shape[0]

                    ss=[]
                    x1=[]
                    cc=[]
                    src_color =(kırmızı, kırmızı, kırmızı)
                    src_color1=(255,255,255)
                    image=cal
                    # Compute matching pixels
                    mask = cv.inRange(image, src_color, src_color1)

                    # Replace pixels where the mask is nonzero
                    if black==0:
                        if kırmızı==0:
                            image[mask > 0] = [255, 255, 255]
                        else:
                            image[mask > 0] = [0, 0, 255]

                    else:
                        if kırmızı==0:
                            image[mask > 0] = [255, 255, 255]
                        else:
                            image[mask < 255] = [0, 0, 255]

                    return image
                kırmızı=14
                brightness=326
                contrast=226
                fotos=int(input("How many images will be processed ?"))
                for i in range(fotos):
                    im1=cv.imread(("{}.jpg").format(i))
                    image1 = im1.copy()
                    effect = pixcont(image1,kırmızı, brightness,
                                						contrast)
                    image=effect.copy()
                    gray = cv.cvtColor(image, cv.COLOR_BGR2GRAY)
                    gray_three = cv.merge([gray,gray,gray])
                    a=gray_three.ndim
                    cv.waitKey(0)
                    wid = image.shape[1]
                    hgt = image.shape[0]
                    Redones=0
                    xlis=[]
                    src_color =(1, 1, 1)
                    src_color1=(255,255,255)
                    mask = cv.inRange(image, src_color, src_color1)
                    #kırmızı olanları beyaza çevir
                    for y in range(0,hgt,1):
                        for x in range (0,wid,1):
                            if (gray_three[y,x]>[0,0,0]).all():
                                gray_three[y,x]=[255,255,255]
                            if (gray_three[y,x]==[0,0,0]).all():
                                gray_three[y,x]=[0,0,0]
                    cv.imwrite(("{}.png").format(i),gray_three)
                    print(i)
                print("finish")

                for i in range(fotos):
                    hello=cv.imread(("{}.png").format(i))
                    gray1 = cv.cvtColor(hello, cv.COLOR_BGR2GRAY)
                    edged = cv.Canny(gray1, 30, 200)
                    cv.waitKey(0)
                    wid = gray1.shape[1]
                    hgt = gray1.shape[0]
                    contours, hierarchy = cv.findContours(edged,
                        cv.RETR_EXTERNAL, cv.CHAIN_APPROX_NONE)
                    len_con=len(contours)
                    azazil=[]
                    for v in range(len_con):
                        azazil.append(cv.arcLength(contours[v],True))
                    ort=sum(azazil)/len(azazil)
                    maksat=max(azazil)
                    draw=[]
                    for ss in range(len(azazil)):
                        if (maksat-ort)<= azazil[ss] <=(maksat+ort):
                            draw.append(ss)
                    for dd in draw:
                        xcorforav=[v[0] for v in contours[dd]]
                        ycorforav=[v[0] for v in contours[dd]]
                    coordxforav=[]
                    coordyforav=[]
                    for cc in range(len(xcorforav)):
                        coordxforav.append(xcorforav[cc])
                        coordyforav.append(ycorforav[cc])
                    xs=[v[0] for v in coordxforav]
                    ys=[v[1] for v in coordyforav]


                    ymin,ymax=min(ys),max(ys)
                    yaverage=(ymax-ymin)
                    ymid=sum(ys)/len(ys)
                    yplus=ymid+yaverage+30
                    yex=ymid-yaverage-30
                    ylastama=[]

                    for ss in range(len(contours)):
                        xcorforadd=[v[0] for v in contours[ss]]
                        ycorforadd=[v[0] for v in contours[ss]]
                        coordxforadd=[]
                        coordyforadd=[]
                        for cc in range(len(xcorforadd)):
                            coordxforadd.append(xcorforadd[cc])
                            coordyforadd.append(ycorforadd[cc])
                        xsadd=[v[0] for v in coordxforadd]
                        ysadd=[v[1] for v in coordyforadd]
                        for ff in range(len(ysadd)):
                            if yex<ysadd[ff]<yplus:
                                draw.append(ss)
                    alan=set(draw)

                    bw = np.zeros((hgt,wid,3),np.uint8)
                    for f in range(len(contours)):
                        if f in draw:
                            cv.drawContours(bw, [contours[f]], 0, (255,255,255), 1)
                #
                #y ortak hizası metodunu dene
                #
                    im = Image.fromarray(bw)
                    im.save("{}.jpeg".format(i))
                    lina_color = io.imread("{}.jpeg".format(i))
                    lina_gray = color.rgb2gray(lina_color)
                    image_array=np.array(lina_gray,dtype=bool)
                    feret_diameter, feret_idx = statistical_length.feret(image_array)
                    max_feret = feret_diameter.max()
                    min_feret = feret_diameter.min()
                    position = (10,50)
                    cv.putText(
                     lina_color,
                     "Thickness:{}".format(max_feret),
                     position,
                     cv.FONT_HERSHEY_SIMPLEX,
                     1,
                     (209, 80, 0, 255),
                     1)
                    cv.imwrite('output{}.png'.format(i), lina_color)
                print("finish")
            except Exception as e:
                print(e)
        vf=ttk.Button(self, text='Volume Fraction', command=volfrac)
        vf.pack()
        CALCULATOESPANYALO=ttk.Button(self,text="Calculator",command=calculator)
        CALCULATOESPANYALO.pack()
        blacklivesmatter=ttk.Checkbutton(self,text="For Black Kind Fibers Please Click the Button",command=blacklivesmatter)
        blacklivesmatter.pack()
        def color_match_1():
            global filename,basename1
            try:
                renk_match_command=img_color_match(filename)
                w_named=os.path.splitext(basename1)[0]
                cv.imshow(w_named,renk_match_command)
                cv.waitKey(0)
                cv.destroyAllWindows()
                return renk_match_command
            except Exception as e:
                print(e)
                #popupmsg("Please upload an image to run the code")
        renk_match=ttk.Button(self,text="Match",command=color_match_1)
        renk_match.pack()
        def brightness():
            global basename1
            try:
                def BrightnessContrast(brightness=0):
                    global effect
                	# getTrackbarPos returns the current
                	# position of the specified trackbar.
                    brightness = cv.getTrackbarPos('Brightness',
								   	             'GEEK')
                    contrast = cv.getTrackbarPos('Contrast',
                								'GEEK')


                    effect = controller(img, brightness,
                						contrast)

                	# The function imshow displays an image
                	# in the specified window
               	    cv.imshow('Brightness and Contrast', effect)

                def controller(img, brightness=255,
                			contrast=127):

                	brightness = int((brightness - 0) * (255 - (-255)) / (510 - 0) + (-255))

                	contrast = int((contrast - 0) * (127 - (-127)) / (254 - 0) + (-127))

                	if brightness != 0:
                		if brightness > 0:
                			shadow = brightness
                			max = 255
                		else:
                			shadow = 0
                			max = 255 + brightness
                		al_pha = (max - shadow) / 255
                		ga_mma = shadow
                		# The function addWeighted calculates
                		# the weighted sum of two arrays
                		cal = cv.addWeighted(img, al_pha,
                							img, 0, ga_mma)
                	else:
                		cal = img
                	if contrast != 0:
                		Alpha = float(131 * (contrast + 127)) / (127 * (131 - contrast))
                		Gamma = 127 * (1 - Alpha)
                		# The function addWeighted calculates
                		# the weighted sum of two arrays
                		cal = cv.addWeighted(cal, Alpha,
                							cal, 0, Gamma)
                	# putText renders the specified text string in the image.
                	cv.putText(original, 'B:{},C:{}'.format(brightness,
                										contrast), (10, 30),
                				cv.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                	return cal
                if __name__ == '__main__':
                	# The function imread loads an image
                	# from the specified file and returns it.
                	original = cv.imread(basename1)
                	# Making another copy of an image.
                	img = original.copy()
                	# The function namedWindow creates a
                	# window that can be used as a placeholder
                	# for images.
                	cv.namedWindow('GEEK')
                	# The function imshow displays an
                	# image in the specified window.
                	cv.imshow('GEEK', original)
                	# createTrackbar(trackbarName,
                	# windowName, value, count, onChange)
                	# Brightness range -255 to 255
                	cv.createTrackbar('Brightness',
                					'GEEK', 255, 2 * 255,
                					BrightnessContrast)
                	# Contrast range -127 to 127
                	cv.createTrackbar('Contrast', 'GEEK',
                					127, 2 * 127,
                					BrightnessContrast)


                	BrightnessContrast(0)

                # The function waitKey waits for
                # a key event infinitely or for delay
                # milliseconds, when it is positive.
                cv.waitKey(0)
                cv.imwrite("cont.png",effect)


            except Exception as e:
                print(e)
                popupmsg(e)

class Mudo(tk.Tk):
    def __init__(self,*args,**kwargs):
        tk.Tk.__init__(self,*args,**kwargs)
        tk.Tk.iconbitmap(self,default="logo.ico")
        tk.Tk.wm_title(self,"VIQUAC")
        global menubar
        global filemenu
        global disable_menu
        global enable_menu
        global disable_menu
        global xl
        global fotoolcer
        global pop_yes_or_no

        def logout():
            print(fotoolcer)
            global disable_menu
            global label1
            if fotoolcer==0:
                x=0
            else:
                label1.pack_forget()

            while True:

                disable_menu()
                command=lambda:self.show_frame(warningPage)
                command()
                break



        container=tk.Frame(self)
        container.pack(side="top",fill="both",expand=True)

        container.grid_rowconfigure(0,weight=1)
        container.grid_columnconfigure(0,weight=1)

        menubar=tk.Menu(container)

        def disable_menu():
            menubar.entryconfig("3", state="disabled")
            menubar.entryconfig("2", state="disabled")
            menubar.entryconfig("4", state="disabled")

        def enable_menu():
            menubar.entryconfig("3", state="normal")
            menubar.entryconfig("2", state="normal")
            menubar.entryconfig("4", state="normal")

        filemenu=tk.Menu(menubar,tearoff=0)
        filemenu.add_command(label="Exit",command=destroyer)
        menubar.add_cascade(label="Exit",menu=filemenu)

##        helpmenu=tk.Menu(menubar,tearoff=0)
##        menubar.add_cascade(label="Help",menu=helpmenu)
##        xlsl=tk.Menu(menubar,tearoff=0)
##        xlsl.add_command(label="Export to .XLSX",command="")
##
##        menubar.add_cascade(label="Export",menu=xlsl)

        logut=tk.Menu(menubar,tearoff=0)
        logut.add_command(label="Log-Out",command=logout)
        menubar.add_cascade(label="Log-Out",menu=logut)


        tk.Tk.config(self,menu=menubar)
        self.frames={}

        for F in [warningPage,registerPage,composite]:
            frame =F(container,self)
            self.frames[F]=frame
            frame.grid(row=0,column=0,sticky="nsew")
        self.show_frame(composite)
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()
def destroyer():
    app.destroy()

def popupmsg(msg):
    popup=tk.Toplevel()
    def leavemini():
        popup.destroy()
    popup.wm_title("VIQUAC")
    label=ttk.Label(popup,text=msg,font=Normal_Font)
    label.pack(side="top",fill="x",pady=10)
    B1=ttk.Button(popup,text="Close",command =leavemini)
    B1.pack()
    popup.wait_window()



app = Mudo()

##app.attributes('-transparent',True)
app.bind("<Escape>",lambda x : app.destroy())
app.geometry("500x250")
app.mainloop()